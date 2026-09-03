from __future__ import annotations

import json
import logging
import shutil
import tempfile
import threading
import uuid
from contextlib import contextmanager
from datetime import date, datetime
from decimal import Decimal
from pathlib import Path
from typing import Any, Iterator
from urllib.parse import unquote, urlparse

import duckdb
import httpx
import sqlglot
from openpyxl import load_workbook
from sqlglot import expressions as exp

logger = logging.getLogger("excel_analysis")


def _json_value(value: Any) -> Any:
    if isinstance(value, (date, datetime)):
        return value.isoformat()
    if isinstance(value, (bytes, bytearray)):
        return bytes(value).hex()
    if isinstance(value, Decimal):
        return str(value)
    return value


def _quote(identifier: str) -> str:
    return '"' + identifier.replace('"', '""') + '"'


def _sql_string(value: str | Path) -> str:
    return "'" + str(value).replace("'", "''") + "'"


def _safe_headers(values: tuple[Any, ...]) -> list[str]:
    result: list[str] = []
    used: set[str] = set()
    for index, value in enumerate(values, 1):
        base = str(value).strip() if value is not None else ""
        base = base or f"column_{index}"
        name, suffix = base, 2
        while name.casefold() in used:
            name = f"{base}_{suffix}"
            suffix += 1
        used.add(name.casefold())
        result.append(name)
    return result


def _filename_from_url(url: str) -> str:
    return Path(unquote(urlparse(url).path)).name or "workbook"


class WorkbookAnalyzer:
    def __init__(self, max_bytes: int = 50 * 1024 * 1024, max_rows: int = 1000,
                 timeout_seconds: int = 30, memory_limit: str = "512MB", storage: Any = None):
        self.max_bytes = max_bytes
        self.max_rows = max_rows
        self.timeout_seconds = timeout_seconds
        self.memory_limit = memory_limit
        self.storage = storage

    def inspect(self, file_url: str) -> dict[str, Any]:
        if not self.storage:
            raise ValueError("persistent storage is required")
        filename = _filename_from_url(file_url)
        logger.info("inspecting workbook filename=%s", filename)
        cache_id = str(uuid.uuid4())
        source_bytes = self._download_bytes(file_url)
        with self._database(file_url=file_url, source_bytes=source_bytes, allow_external_access=True) as (connection, filename, file_format):
            sheets = []
            for table in self._tables(connection):
                columns = [
                    {"name": row[0], "type": row[1]}
                    for row in connection.execute(f"DESCRIBE {_quote(table)}").fetchall()
                ]
                row_count = connection.execute(f"SELECT COUNT(*) FROM {_quote(table)}").fetchone()[0]
                sample = connection.execute(
                    f"SELECT * FROM {_quote(table)} LIMIT 10"
                ).fetchall()
                sheets.append({
                    "name": table,
                    "row_count": row_count,
                    "columns": columns,
                    "sample_rows": [[_json_value(value) for value in row] for row in sample],
                })
            parquet_sources = []
            for index, table in enumerate(self._tables(connection)):
                key = f"parquet:{cache_id}:{index}"
                parquet_path = Path(tempfile.mkdtemp(prefix="dify-parquet-")) / f"{index}.parquet"
                try:
                    connection.execute(
                        f"COPY (SELECT * FROM {_quote(table)}) TO {_sql_string(parquet_path)} (FORMAT PARQUET)"
                    )
                    self.storage.set(key, parquet_path.read_bytes())
                    parquet_sources.append({"name": table, "key": key})
                finally:
                    shutil.rmtree(parquet_path.parent, ignore_errors=True)
            self.storage.set(f"meta:{cache_id}", json.dumps({"filename": filename, "format": file_format, "sheets": parquet_sources}).encode("utf-8"))
            result = {
                "file_url": file_url,
                "cache_id": cache_id,
                "filename": filename,
                "format": file_format,
                "sheets": sheets,
            }
            logger.info("workbook inspection completed cache_id=%s sheets=%d", cache_id, len(sheets))
            return result

    def query(self, cache_id: str, sql: str, limit: int = 1000) -> dict[str, Any]:
        if not self.storage:
            raise ValueError("persistent storage is required")
        try:
            meta = json.loads(self.storage.get(f"meta:{cache_id}").decode("utf-8"))
            parquet_sources = []
            for sheet in meta["sheets"]:
                parquet_sources.append((sheet["name"], self.storage.get(sheet["key"])))
        except Exception as exc:
            raise ValueError(f"cache_id {cache_id} not found") from exc
        if not parquet_sources:
            raise ValueError(f"cache_id {cache_id} not found")
        statements = sqlglot.parse(sql, read="duckdb")
        if not statements or any(not isinstance(statement, exp.Query) for statement in statements):
            raise ValueError("only read-only SELECT or CTE statements are allowed")
        forbidden = {"Insert", "Update", "Delete", "Create", "Drop", "Alter", "Command", "Copy", "Attach", "Merge"}
        for statement in statements:
            if any(node.__class__.__name__ in forbidden for node in statement.walk()):
                raise ValueError("SQL statement is not read-only")
        effective_limit = self._limit(limit)
        logger.info("executing workbook query cache_id=%s statements=%d limit=%d", cache_id, len(statements), effective_limit)
        with self._database(parquet_sources=parquet_sources) as (connection, _filename, _format):
            results = []
            for statement in statements:
                normalized = statement.sql(dialect="duckdb")
                timer = threading.Timer(self.timeout_seconds, connection.interrupt)
                try:
                    timer.start()
                    cursor = connection.execute(
                        f"SELECT * FROM ({normalized}) AS _result LIMIT ?",
                        [effective_limit + 1],
                    )
                    columns = [{"name": item[0], "type": str(item[1])} for item in cursor.description]
                    rows = [[_json_value(value) for value in row] for row in cursor.fetchall()]
                except duckdb.InterruptException as exc:
                    raise TimeoutError(f"query exceeded {self.timeout_seconds} seconds") from exc
                finally:
                    timer.cancel()
                truncated = len(rows) > effective_limit
                if truncated:
                    rows = rows[:effective_limit]
                results.append({"columns": columns, "rows": rows, "row_count": len(rows), "truncated": truncated})
            logger.info("workbook query completed cache_id=%s result_sets=%d", cache_id, len(results))
            return {"cache_id": cache_id, "results": results}

    @contextmanager
    def _database(self, file_url: str | None = None, source_bytes: bytes | None = None,
                  parquet_sources: list[tuple[str, bytes]] | None = None,
                  allow_external_access: bool = False) -> Iterator[tuple[duckdb.DuckDBPyConnection, str, str]]:
        if not file_url and source_bytes is None and parquet_sources is None:
            raise ValueError("either file_url, source_bytes, or parquet_sources is required")
        filename = _filename_from_url(file_url) if file_url else "workbook"
        with tempfile.TemporaryDirectory(prefix="dify-excel-") as directory:
            if parquet_sources is not None:
                database = Path(directory) / "workbook.duckdb"
                connection = duckdb.connect(str(database), config={"memory_limit": self.memory_limit})
                try:
                    for index, (table, data) in enumerate(parquet_sources):
                        path = Path(directory) / f"{index}.parquet"
                        path.write_bytes(data)
                        connection.execute(
                            f"CREATE TEMP TABLE {_quote(table)} AS SELECT * FROM read_parquet({_sql_string(path)})"
                        )
                    if not allow_external_access:
                        connection.execute("SET enable_external_access=false")
                    yield connection, filename, "parquet"
                finally:
                    connection.close()
                return
            source = Path(directory) / "source"
            if source_bytes is None:
                self._download(file_url, source)
            else:
                source.write_bytes(source_bytes)
            connection = None
            selected = None
            error: Exception | None = None
            # The Dify preview URL often has no real filename suffix. Detect the
            # format by importing the downloaded bytes, not by inspecting its URL.
            for suffix in (".xlsx", ".csv"):
                database = Path(directory) / "workbook.duckdb"
                try:
                    connection = duckdb.connect(
                        str(database),
                        config={"memory_limit": self.memory_limit},
                    )
                    if suffix == ".csv":
                        connection.execute('CREATE TABLE "data" AS SELECT * FROM read_csv_auto(?)', [str(source)])
                    else:
                        self._import_xlsx(connection, source)
                    # CSV/XLSX import needs local file access; user SQL must not.
                    if not allow_external_access:
                        connection.execute("SET enable_external_access=false")
                    selected = suffix
                    break
                except Exception as exc:
                    error = exc
                    if connection:
                        connection.close()
                    connection = None
                    database.unlink(missing_ok=True)
            if connection is None or selected is None:
                raise ValueError("file cannot be read as .xlsx or .csv") from error
            try:
                yield connection, filename, selected[1:]
            finally:
                connection.close()

    def _download_bytes(self, url: str) -> bytes:
        with tempfile.TemporaryDirectory(prefix="dify-excel-download-") as directory:
            destination = Path(directory) / "source"
            self._download(url, destination)
            return destination.read_bytes()

    def _download(self, url: str, destination: Path) -> None:
        current = url
        with httpx.Client(timeout=httpx.Timeout(30), follow_redirects=False) as client:
            for redirect_count in range(6):
                logger.debug("downloading workbook redirect=%d", redirect_count)
                with client.stream("GET", current, headers={"User-Agent": "dify-excel-plugin/1.0"}) as response:
                    if response.is_redirect:
                        location = response.headers.get("location")
                        if not location:
                            raise ValueError("download redirect has no location")
                        current = str(response.url.join(location))
                        continue
                    response.raise_for_status()
                    content_length = response.headers.get("content-length")
                    if content_length and int(content_length) > self.max_bytes:
                        raise ValueError("workbook exceeds the configured size limit")
                    size = 0
                    with destination.open("wb") as output:
                        for chunk in response.iter_bytes():
                            size += len(chunk)
                            if size > self.max_bytes:
                                raise ValueError("workbook exceeds the configured size limit")
                            output.write(chunk)
                    logger.info("workbook download completed bytes=%d", size)
                    return
            raise ValueError("too many download redirects")

    def _import_xlsx(self, connection: duckdb.DuckDBPyConnection, source: Path) -> None:
        # openpyxl validates the filename suffix before inspecting the file bytes.
        workbook_source = source.with_suffix(".xlsx")
        shutil.copyfile(source, workbook_source)
        workbook = load_workbook(workbook_source, read_only=True, data_only=True)
        try:
            for sheet in workbook.worksheets:
                rows = sheet.iter_rows(values_only=True)
                try:
                    headers = _safe_headers(next(rows))
                except StopIteration:
                    headers = ["column_1"]
                    rows = iter(())
                connection.execute(
                    f"CREATE TABLE {_quote(sheet.title)} ({', '.join(f'{_quote(header)} VARCHAR' for header in headers)})"
                )
                values = [tuple(row[:len(headers)]) for row in rows]
                if values:
                    connection.executemany(
                        f"INSERT INTO {_quote(sheet.title)} VALUES ({', '.join('?' for _ in headers)})",
                        values,
                    )
        finally:
            workbook.close()
            workbook_source.unlink(missing_ok=True)

    @staticmethod
    def _tables(connection: duckdb.DuckDBPyConnection) -> list[str]:
        return [row[0] for row in connection.execute("SHOW TABLES").fetchall()]

    def _limit(self, value: int) -> int:
        if value < 1:
            raise ValueError("limit must be positive")
        return min(value, self.max_rows)


def analyzer_from_env(storage: Any = None) -> WorkbookAnalyzer:
    import os
    return WorkbookAnalyzer(
        max_bytes=int(os.getenv("EXCEL_MAX_BYTES", str(50 * 1024 * 1024))),
        max_rows=int(os.getenv("EXCEL_SQL_MAX_ROWS", "1000")),
        timeout_seconds=int(os.getenv("EXCEL_SQL_TIMEOUT_SECONDS", "30")),
        memory_limit=os.getenv("EXCEL_SQL_MEMORY_LIMIT", "512MB"),
        storage=storage,
    )
